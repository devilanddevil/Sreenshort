import os
import re
import datetime
import io
from flask import Flask, render_template, request, redirect, url_for, flash, Response, abort, send_file
import pytesseract
from PIL import Image, ImageOps 
from werkzeug.utils import secure_filename
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func
import uuid
import cloudinary
import cloudinary.uploader
import cloudinary.api

# Initialize Flask App
app = Flask(__name__)
app.secret_key = 'super_secret_key_for_flash_messages'

# --- DATABASE CONFIGURATION ---
db_url = os.environ.get('DATABASE_URL', 'sqlite:///idle_tracker.db')
if db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql://", 1)
app.config['SQLALCHEMY_DATABASE_URI'] = db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

app.config['UPLOAD_FOLDER'] = 'static/uploads'

# Initialize Extensions
db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

# --- MODELS ---
class User(UserMixin, db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(150), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    is_admin = db.Column(db.Boolean, default=False)
    records = db.relationship('Record', backref='user', lazy=True, cascade="all, delete-orphan")

class Record(db.Model):
    __tablename__ = 'records'
    id = db.Column(db.Integer, primary_key=True)
    timestamp = db.Column(db.DateTime, default=datetime.datetime.now)
    filename = db.Column(db.String(300))
    idle_minutes = db.Column(db.Integer)
    original_text = db.Column(db.Text)
    reason = db.Column(db.String(300), nullable=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# --- TESSERACT SETUP ---
tesseract_paths = [
    r'C:\Program Files\Tesseract-OCR\tesseract.exe',
    os.path.join(os.environ.get('LOCALAPPDATA', ''), r'Tesseract-OCR\tesseract.exe'),
    r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe'
]

tesseract_cmd = None
for path in tesseract_paths:
    if os.path.exists(path):
        tesseract_cmd = path
        break


if tesseract_cmd:
    pytesseract.pytesseract.tesseract_cmd = tesseract_cmd
    print(f"Using Tesseract at: {tesseract_cmd}")
else:
    print("Tesseract not found in common paths. Relying on system PATH.")

# Configure Cloudinary
cloudinary.config(
    cloud_name = os.environ.get('CLOUD_NAME'),
    api_key = os.environ.get('API_KEY'),
    api_secret = os.environ.get('API_SECRET'),
    secure = True
)

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# --- HELPER FUNCTIONS ---
def parse_idle_time(text):
    print(f"DEBUG OCR TEXT:\n{text}\n----------------") 
    patterns = [
        r'idle for:?\s*.*?((\d+\s*h\s*)?(\d+\s*m))',
        r'Not Working\s*[-–—]?\s*.*?((\d+\s*h\s*)?(\d+\s*m))'
    ]
    time_str = None
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        if match:
            time_str = match.group(1).lower()
            break
            
    if time_str:
        hours = 0
        minutes = 0
        h_match = re.search(r'(\d+)\s*h', time_str)
        if h_match: hours = int(h_match.group(1))
        m_match = re.search(r'(\d+)\s*m', time_str)
        if m_match: minutes = int(m_match.group(1))
        total_minutes = (hours * 60) + minutes
        if total_minutes > 0:
            return total_minutes, time_str.strip()
    return 0, "No time detected"

def get_stats(page=1, per_page=5, user_id=None):
    if not user_id:
        return {'today': 0, 'month': 0, 'recent': {'records': [], 'page': 1, 'per_page': 5, 'total_pages': 0, 'total_count': 0}}
    
    now = datetime.datetime.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    today_minutes = db.session.query(func.sum(Record.idle_minutes)).filter(
        Record.user_id == user_id, 
        Record.timestamp >= today_start
    ).scalar() or 0
    
    month_minutes = db.session.query(func.sum(Record.idle_minutes)).filter(
        Record.user_id == user_id, 
        Record.timestamp >= month_start
    ).scalar() or 0
    
    pagination = Record.query.filter_by(user_id=user_id).order_by(Record.timestamp.desc()).paginate(page=page, per_page=per_page, error_out=False)
    
    return {
        'today': today_minutes,
        'month': month_minutes,
        'recent': {
            'records': pagination.items,
            'page': page,
            'per_page': per_page,
            'total_pages': pagination.pages,
            'total_count': pagination.total
        }
    }

def format_minutes(total_minutes):
    if total_minutes is None: return "0m"
    h = total_minutes // 60
    m = total_minutes % 60
    return f"{h}h {m}m" if h > 0 else f"{m}m"

app.jinja_env.filters['format_minutes'] = format_minutes

# --- ROUTES ---

@app.route('/emergency_reset')
def emergency_reset():
    # Force reset ID 1 to admin/12345
    try:
        user = User.query.get(1)
        if user:
            user.username = 'admin'
            user.password_hash = generate_password_hash('12345')
            user.is_admin = True
            db.session.commit()
            return "SUCCESS: User ID 1 reset to username='admin', password='12345'. <a href='/login'>Go to Login</a>"
        else:
            return "ERROR: User ID 1 not found. Please Register a new user first."
    except Exception as e:
        return f"ERROR: {str(e)}"

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
        
    if request.method == 'POST':
        username = request.form['username'].strip() # Strip whitespace
        password = request.form['password']
        
        # Case insensitive login
        user = User.query.filter(func.lower(User.username) == func.lower(username)).first()
        
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            return redirect(url_for('index'))
        else:
            flash('Invalid username or password', 'error')
            
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
        
    if request.method == 'POST':
        username = request.form['username'].strip() # Strip whitespace
        password = request.form['password']
        confirm_password = request.form['confirm_password']
        
        if password != confirm_password:
            flash('Passwords do not match', 'error')
            return redirect(url_for('register'))
            
        if User.query.filter(func.lower(User.username) == func.lower(username)).first():
            flash('Username already exists', 'error')
            return redirect(url_for('register'))
            
        user_count = User.query.count()
        is_admin = (user_count == 0)
        
        hashed_pw = generate_password_hash(password)
        new_user = User(username=username, password_hash=hashed_pw, is_admin=is_admin)
        db.session.add(new_user)
        db.session.commit()
        
        login_user(new_user)
        flash('Account created successfully!', 'success')
        return redirect(url_for('index'))
    return render_template('register.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    stats = get_stats(page=1, per_page=5, user_id=current_user.id)
    return render_template('index.html', stats=stats)

@app.route('/upload', methods=['POST'])
@login_required
def upload_file():
    manual_minutes = request.form.get('manual_minutes')
    reason = request.form.get('reason', '')
    
    if manual_minutes:
        try:
            minutes = int(manual_minutes)
            if minutes > 0:
                record = Record(
                    filename='manual_entry',
                    idle_minutes=minutes,
                    original_text="Manual Entry",
                    reason=reason,
                    user_id=current_user.id,
                    timestamp=datetime.datetime.now()
                )
                db.session.add(record)
                db.session.commit()
                flash(f'Success! Manually logged {format_minutes(minutes)}.', 'success')
            else:
                flash('Minutes must be positive.', 'error')
        except ValueError:
            flash('Invalid minutes value.', 'error')
        return redirect(url_for('index'))

    if 'screenshot' not in request.files:
        flash('No file part and no manual minutes provided.')
        return redirect(request.url)
    
    file = request.files['screenshot']
    
    if file.filename == '' and not manual_minutes:
        flash('No selected file and no manual minutes.')
        return redirect(request.url)
    
    if file:
        original_filename = secure_filename(file.filename)
        # Generate unique filename to prevent overwriting
        timestamp = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
        unique_id = str(uuid.uuid4())[:8]
        filename = f"{timestamp}_{unique_id}_{original_filename}"
        
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        try:
            image_bytes = file.read()
            if len(image_bytes) == 0:
                flash('Uploaded file is empty.', 'error')
                return redirect(request.url)
                
            image_stream = io.BytesIO(image_bytes)
            base_img = Image.open(image_stream)
            
            # Upload to Cloudinary
            upload_result = cloudinary.uploader.upload(
                io.BytesIO(image_bytes), 
                public_id=filename.split('.')[0],
                resource_type="image"
            )
            file_url_or_id = upload_result.get('secure_url')
            
            img_gray = base_img.convert('L')
            text = pytesseract.image_to_string(img_gray)
            minutes, detected_text = parse_idle_time(text)
            
            if minutes == 0:
                text = pytesseract.image_to_string(img_gray, config='--psm 6')
                minutes, detected_text = parse_idle_time(text)

            if minutes == 0:
                try:
                    img_thresh = img_gray.point(lambda p: p > 150 and 255)
                    text = pytesseract.image_to_string(img_thresh, config='--psm 6')
                    minutes, detected_text = parse_idle_time(text)
                except: pass

            if minutes == 0:
                 try:
                     img_invert = ImageOps.invert(base_img.convert('RGB')).convert('L')
                     text = pytesseract.image_to_string(img_invert, config='--psm 6')
                     minutes, detected_text = parse_idle_time(text)
                 except: pass

            if minutes > 0:
                record = Record(
                    filename=file_url_or_id,
                    idle_minutes=minutes,
                    original_text=detected_text,
                    reason=reason,
                    user_id=current_user.id,
                    timestamp=datetime.datetime.now()
                )
                db.session.add(record)
                db.session.commit()
                flash(f'Success! Logged {format_minutes(minutes)} of idle time.', 'success')
            else:
                flash(f'Could not detect idle time. OCR Text: "{text.strip()[:100]}..."', 'error')
                
        except Exception as e:
            flash(f'Error processing image: {str(e)}', 'error')
            
        return redirect(url_for('index'))

import openpyxl
from openpyxl.styles import Font

@app.route('/history')
@login_required
def history():
    month_filter = request.args.get('month')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 5, type=int)
    
    if per_page > 100: per_page = 100
    if per_page < 1: per_page = 5
    
    target_user_id = current_user.id
    target_username = current_user.username
    
    if current_user.is_admin and request.args.get('user_id'):
        target_user_id = request.args.get('user_id', type=int)
        u = User.query.get(target_user_id)
        if u: target_username = u.username
    
    # 1. Base Filter for Records (Pagination)
    query = Record.query.filter_by(user_id=target_user_id)
    
    dates = db.session.query(Record.timestamp).filter_by(user_id=target_user_id).all()
    available_months = sorted(list(set([d[0].strftime('%Y-%m') for d in dates])), reverse=True)
    
    if month_filter:
        try:
             y, m = map(int, month_filter.split('-'))
             start_date = datetime.datetime(y, m, 1)
             if m == 12: end_date = datetime.datetime(y + 1, 1, 1)
             else: end_date = datetime.datetime(y, m + 1, 1)
             query = query.filter(Record.timestamp >= start_date, Record.timestamp < end_date)
        except: pass

    pagination = query.order_by(Record.timestamp.desc()).paginate(page=page, per_page=per_page, error_out=False)
    records = pagination.items
    
    # 2. SEPARATE Sum Query for Total Minutes (Fixes Aggregation Issue)
    sum_query = db.session.query(func.sum(Record.idle_minutes)).filter(Record.user_id == target_user_id)
    
    # Apply same date filter to sum query
    if month_filter:
        try:
             y, m = map(int, month_filter.split('-'))
             start_date = datetime.datetime(y, m, 1)
             if m == 12: end_date = datetime.datetime(y + 1, 1, 1)
             else: end_date = datetime.datetime(y, m + 1, 1)
             sum_query = sum_query.filter(Record.timestamp >= start_date, Record.timestamp < end_date)
        except: pass

    total_minutes = sum_query.scalar() or 0
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.args.get('ajax') == '1':
        return render_template('partials/table_content.html', 
                               records=records,
                               page=page,
                               per_page=per_page,
                               total_pages=pagination.pages,
                               endpoint='history',
                               current_month=month_filter,
                               target_user_id=target_user_id)
    
    return render_template('history.html', 
                           records=records, 
                           available_months=available_months, 
                           current_month=month_filter, 
                           total_minutes=total_minutes,
                           page=page,
                           per_page=per_page,
                           total_pages=pagination.pages,
                           target_username=target_username if target_user_id != current_user.id else None,
                           target_user_id=target_user_id,
                           endpoint='history')

@app.route('/admin')
@login_required
def admin():
    if not current_user.is_admin: abort(403)
    
    users = db.session.query(
        User,
        func.sum(Record.idle_minutes).label('total_minutes'),
        func.count(Record.id).label('total_records')
    ).outerjoin(Record).group_by(User.id).all()
    
    user_list = []
    for u, tm, tr in users:
        user_list.append({
            'id': u.id,
            'username': u.username,
            'is_admin': u.is_admin,
            'total_minutes': tm or 0,
            'total_records': tr or 0
        })
        
    return render_template('admin.html', users=user_list)

@app.route('/admin/edit_user/<int:user_id>', methods=['GET', 'POST'])
@login_required
def edit_user(user_id):
    if not current_user.is_admin: abort(403)
    user = User.query.get_or_404(user_id)
    
    if request.method == 'POST':
        new_username = request.form['username'].strip()
        existing = User.query.filter(func.lower(User.username) == func.lower(new_username)).first()
        if existing and existing.id != user.id:
            flash('Username already exists.', 'error')
        else:
            user.username = new_username
            db.session.commit()
            flash(f'Username updated to {new_username}.', 'success')
            return redirect(url_for('admin'))
            
    return render_template('edit_user.html', user=user)

@app.route('/admin/delete_user/<int:user_id>', methods=['POST'])
@login_required
def delete_user(user_id):
    if not current_user.is_admin: abort(403)
    if user_id == current_user.id:
        flash('You cannot delete your own account!', 'error')
        return redirect(url_for('admin'))
    
    user = User.query.get(user_id)
    if user:
        db.session.delete(user)
        db.session.commit()
        flash('User deleted.', 'success')
    return redirect(url_for('admin'))

@app.route('/admin/reset_password/<int:user_id>', methods=['GET', 'POST'])
@login_required
def reset_password(user_id):
    if not current_user.is_admin: abort(403)
    user = User.query.get_or_404(user_id)
    
    if request.method == 'POST':
        new_password = request.form['new_password']
        user.password_hash = generate_password_hash(new_password)
        db.session.commit()
        flash(f'Password for {user.username} reset.', 'success')
        return redirect(url_for('admin'))
        
    return render_template('reset_password.html', user_id=user.id, username=user.username)

@app.route('/export_excel')
@login_required
def export_excel():
    records = Record.query.filter_by(user_id=current_user.id).order_by(Record.timestamp.desc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Idle Time Report"
    
    headers = ['ID', 'Date', 'Idle Time', 'Reason', 'Detected Text', 'Screenshot Link']
    ws.append(headers)
    for cell in ws[1]: cell.font = Font(bold=True)
        
    for r in records:
        readable_time = format_minutes(r.idle_minutes)
        # Check if filename is a URL or a local path (legacy support)
        if r.filename.startswith('http'):
            link = r.filename
        else:
             # Legacy or fallback
             link = url_for('static', filename=f'uploads/{r.filename}', _external=True)

        ws.append([r.id, r.timestamp, readable_time, r.reason, r.original_text])
        
        cell = ws.cell(row=ws.max_row, column=6)
        cell.value = "View Screenshot"
        cell.hyperlink = link
        cell.style = "Hyperlink"
    
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 40
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='idle_time_report.xlsx')

@app.route('/delete/<int:record_id>', methods=['POST'])
@login_required
def delete_record(record_id):
    record = Record.query.get_or_404(record_id)
    if record.user_id != current_user.id: abort(403)
    db.session.delete(record)
    db.session.commit()
    flash('Record deleted.', 'info')
    return redirect(url_for('index'))

with app.app_context():
    db.create_all()

if __name__ == '__main__':
    app.run(debug=True)
